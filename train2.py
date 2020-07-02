import numpy as np
import cv2
import xlsxwriter
from playsound import playsound
import datetime
import pandas
from openpyxl import load_workbook
from tkinter import *
def add(k,p,Id):
    file="new_stat.xlsx"
    wb=load_workbook(file)
    ws=wb.worksheets[0]
    ws_tables=[]
    now = datetime.datetime.now()
    ws['A1']='roll'
    ws['B1']='name'
    ws['C1']='Present'
    ws['D1']='Date'
    ws['E1']='Time'

    ws['A'+str(p+1)]=k
    ws['B'+str(p+1)]=str(id)
    ws['C'+str(p+1)]='YES'
    ws['D'+str(p+1)]=now.strftime("%Y-%m-%d")
    ws['E'+str(p+1)]=now.strftime("%H:%M:%S")
    wb.save(file)
    for table in ws_tables:
        ws_tables.append(table)
    #workbook.close()
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
cap = cv2.VideoCapture(0)
rec =cv2.face.LBPHFaceRecognizer_create()
rec.read("trainingdata.yml")
id=0
font = cv2.FONT_HERSHEY_SIMPLEX
#font=cv2.cv.InitFont(cv2.cv.CV_FONT_HERSHEY_COMPLEX_SMALL,5,1,0,4)
while 1:
    ret, img = cap.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.5, 5)
    for (x,y,w,h) in faces:
        cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
        id,conf=rec.predict(gray[y:y+h,x:x+w])
        if(id==2):
            p=id
            k=id
            id="Sourav"
            flag=1
        elif id==1:
            p=id
            k=id
            flag=1
            id="Arnab"
        elif id==3:
            p=id
            k=id
            flag=1
            id="Tutu"
        elif id==4:
            p=id
            k=id
            flag=1
            id="babai"
        elif id==5:
            p=id
            k=id
            flag=1
            id='Mantu'
        elif id==6:
            p=id
            k=id
            id="riju"
            flag=1
        elif id==7:
            p=id
            k=id
            id="Subhojit"
            flag=1
        cv2.putText(img,id, (x,y+h),font,1,(255,255,255))
        add(k,p,id)
        #cv2.cv.puttext(cv2.cv.fromarray(img),str(id),(x,y+h),font,255)
    cv2.imshow('img',img)
    
    if cv2.waitKey(1) == ord('q'):
        break
playsound('sound.mp3')

cap.release()

cv2.destroyAllWindows()


